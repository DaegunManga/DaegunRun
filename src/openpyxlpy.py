from multiprocessing.sharedctypes import Value
import openpyxl
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt

class load_xl :
    def __init__ (self) : 
        self.wb = openpyxl.load_workbook('./UserInfo/score.xlsx')
        self.sheet = self.wb.active

    def save_value (self, id_t, score) :
        self.sheet.cell(row = id_t,column = 1).value = id_t
        self.sheet.cell(row = id_t,column = 2).value = score    
        self.wb.save('./UserInfo/score.xlsx')
        self.wb.close()
        
    def id_return (self) :
        return self.sheet.max_row

    
    def save_Data_IMG (self) :
        self.df = pd.read_excel("./UserInfo/score.xlsx", engine = "openpyxl")
        self.df.loc[len(self.df) + 1] = [None, None]
        self.df = self.df.shift(1)
        self.df.loc[0] = self.df.columns 
        self.df.columns = ['Id', 'Score']
        self.df = self.df.astype({'Id' : 'int'})
        self.df = self.df.astype({'Id' : 'str', 'Score' : 'float'})
        df_sorted_by_values = self.df.sort_values(by='Score' ,ascending=True)
        sns.barplot(x='Id',y='Score', data = df_sorted_by_values.head(5), palette = 'Blues_r')
        plt.savefig('./img/DATA_IMG.png')

            
